from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from home.models import Place

class Command(BaseCommand):
    help = 'Load data from myuni data into Django model'

    def add_arguments(self, parser):
        parser.add_argument('file_path', type=str, help='myuni-data.xlsx')

    def handle(self, *args, **options):
        file_path = options['file_path']
        wb = load_workbook(file_path)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            Place.objects.create(
                id=row[0],
                name=row[1],
                latitude=row[2],
                longitude=row[3],
            )

        self.stdout.write(self.style.SUCCESS('Data loaded successfully'))
